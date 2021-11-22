import openpyxl

#read sheet
wb=openpyxl.load_workbook("Sales.xlsx")
ws = wb.worksheets[0]
cells=ws["A2":"B4"]

print("売上：")
for cell in cells:
    name,sale=cell
    print("{}:{}、{}:{}".format(name.coordinate,name.value,sale.coordinate,sale.value))

address=input("修正する箇所：")
sam_B2=ws[address]
sam_B2.value=int(input("値："))

print("---修正後：---")

print("売上：")
for cell in cells:
    name,sale=cell
    print("{}:{}、{}:{}".format(name.coordinate,name.value,sale.coordinate,sale.value))
