import openpyxl

def get_sheet(wb):
    selec=int(input("which sheet 1~{}:".format(len(wb.worksheets))))
    return wb.worksheets[selec-1]

def get_one_cell(ws):
    select_cell=input("Which the cell you want to get:")
    return ws[select_cell]
    
def get_multiple_cells(ws):
    range_cell=input("Enter the range:")
    return ws[range_cell]

def get_specify_row(ws):
    row=int(input("Enter row:"))
    return ws[row]

func_dict={0:get_one_cell,
          1:get_multiple_cells,
          2:get_specify_row}

#read sheet
wb=openpyxl.load_workbook("Sales.xlsx")
print("--読込完了--")
print(wb.worksheets)

#get sheet
ws=get_sheet(wb)
print("Got sheet:",ws)

#get cell
print("0:Get one cell , 1: Get multiple cells , 2: Specify a row")
selec = int(input("Enter:"))
cell = func_dict[selec](ws)
print(cell)
"""
if selec==0:
    cell=get_one_cell(ws)
    print(cell)
elif selec==1:
    cell=get_multiple_cells(ws)
    print(cell)
elif selec==2:
    cell=get_specify_row(ws)
    print(cell)
"""
