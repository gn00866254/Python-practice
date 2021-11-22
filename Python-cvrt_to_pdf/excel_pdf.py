import os
import win32com.client
import glob
import openpyxl


#抓資料夾底下EXCEL檔名
file=glob.glob("*.xlsx")
#抓目前的資料夾PATH
base = os.path.dirname(__file__)
#結合當前資料夾的PATH和EXCEL檔名
input_file = os.path.join(base, file[0])
#啟動excel
app = win32com.client.Dispatch("Excel.Application")
#讓excel可看見
app.Visible = False
#使確認方塊消失
app.DisplayAlerts = False

#枚数を取得
book = openpyxl.load_workbook(input_file)
sheetnums=len((book.sheetnames))
#取得每張sheet的名字--＞list
sheetALLnames=book.sheetnames

wb = app.Workbooks.Open(input_file)

for i in range(sheetnums):
    wb.Worksheets(i+1).Select()
    xlTypePDF = 0
    app.ActiveSheet.ExportAsFixedFormat(xlTypePDF,
                                        base+"/MEP 受講状況報告書"+sheetALLnames[i]+".pdf")
wb.Close()